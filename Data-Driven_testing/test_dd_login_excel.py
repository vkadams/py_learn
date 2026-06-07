import openpyxl
import pytest
from playwright.sync_api import Page, expect
'''
openpyxl== is required to work with excel files-- pip install openpyxl
'''

# reading data from excel file
login_data = []
workbook = openpyxl.load_workbook(r"C:\Users\61015706\OneDrive - LTIMindtree\Documents\PY\PPY\test_data\data.xlsx")
sheet = workbook.active # workbook["sheetname"]

# reading data from excel sheet
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True): # row is 2 as we dont need to read header
    email, password, validity = row
    login_data.append((str(email or ""), str(password or ""), str(validity or "")))
workbook.close()

# for reading data from json we use for loop
@pytest.mark.parametrize("email, password, validity", login_data)
def test_login_json_datadriven(email, password, validity, page: Page):
    # login form
    page.goto("https://demowebshop.tricentis.com/login")
    page.locator('#Email').fill(email)
    page.locator('#Password').fill(password)
    page.locator('input[value="Log in"]').click()

    # validation
    if validity == "valid":
        expect(page.locator('.ico-logout')).to_be_visible(timeout=5000)  # logout link
    else:
        expect(page.locator(".validation-summary-errors")).to_be_visible(timeout=5000)
        expect(page).to_have_url("https://demowebshop.tricentis.com/login")